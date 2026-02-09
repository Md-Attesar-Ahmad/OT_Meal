import streamlit as st
from openpyxl import load_workbook
from datetime import date, datetime
import math
import os
import re
import pandas as pd

# -----------------------------
# CONFIG
# -----------------------------
FILE_PATH = "OT_Tracker.xlsx"
SHEET_NAME = "OT"

BILLS_DIR = "Bills"
BILLS_INDEX = os.path.join(BILLS_DIR, "bills_index.csv")

st.set_page_config(page_title="OT Meal Tracker", layout="wide")
st.title("OT Meal Tracker")

# -----------------------------
# HELPERS
# -----------------------------
def _to_date(x):
    """Fast conversion to python date (handles datetime/date/strings)."""
    if x is None:
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    try:
        return datetime.fromisoformat(str(x)).date()
    except Exception:
        try:
            return datetime.strptime(str(x), "%d-%b-%Y").date()
        except Exception:
            return None

@st.cache_data(show_spinner=False)
def get_name_to_col(file_path: str, sheet_name: str):
    """Cache header mapping for speed."""
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    mapping = {}
    for col in range(2, ws.max_column + 1):
        name = ws.cell(row=1, column=col).value
        if name:
            mapping[str(name).strip()] = col
    wb.close()
    return mapping

@st.cache_data(show_spinner=False)
def get_all_names(file_path: str, sheet_name: str):
    """Return sorted list of names from header row (B1..)."""
    mapping = get_name_to_col(file_path, sheet_name)
    return sorted(mapping.keys(), key=lambda x: x.lower())

def find_row_for_date(ws, selected_date: date):
    """Find the row index in column A matching selected_date."""
    for idx, (cell_val,) in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=True),
        start=2,
    ):
        if _to_date(cell_val) == selected_date:
            return idx
    return None

def sanitize_filename(s: str) -> str:
    s = str(s).strip()
    s = re.sub(r"[^\w\-. ]+", "_", s)   # keep letters/numbers/_ - . space
    s = re.sub(r"\s+", "_", s)
    return s[:150]

def ensure_bills_storage():
    os.makedirs(BILLS_DIR, exist_ok=True)
    if not os.path.exists(BILLS_INDEX):
        df = pd.DataFrame(columns=["ot_date", "user_name", "file_name", "stored_path", "uploaded_at"])
        df.to_csv(BILLS_INDEX, index=False)

def read_bills_index() -> pd.DataFrame:
    ensure_bills_storage()
    try:
        return pd.read_csv(BILLS_INDEX)
    except Exception:
        return pd.DataFrame(columns=["ot_date", "user_name", "file_name", "stored_path", "uploaded_at"])

def append_bill_index(ot_date: date, user_name: str, file_name: str, stored_path: str):
    df = read_bills_index()
    new_row = {
        "ot_date": ot_date.isoformat(),
        "user_name": user_name,
        "file_name": file_name,
        "stored_path": stored_path.replace("\\", "/"),
        "uploaded_at": datetime.now().isoformat(timespec="seconds"),
    }
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df.to_csv(BILLS_INDEX, index=False)

# -----------------------------
# TAB RENDERERS (IMPORTANT: use return, NOT st.stop)
# -----------------------------
def render_ot_tab():
    st.subheader("OT Meal Tracker")

    selected_date = st.date_input("Select OT Date", value=date.today(), key="ot_date")
    bill_amount = st.number_input("Enter Bill Amount (₹)", step=50, min_value=0, key="ot_bill_amount")

    if bill_amount <= 0:
        st.info("Please enter a bill amount to continue (other tabs will still work).")
        return

    required_people = math.ceil(bill_amount / 750)

    # Load workbook (write mode because we may update)
    try:
        wb = load_workbook(FILE_PATH)
    except Exception as e:
        st.error(f"Unable to open {FILE_PATH}: {e}")
        return

    if SHEET_NAME not in wb.sheetnames:
        st.error(f"Sheet '{SHEET_NAME}' not found in {FILE_PATH}")
        wb.close()
        return

    ws = wb[SHEET_NAME]

    row_index = find_row_for_date(ws, selected_date)
    if row_index is None:
        st.error("Date not found in Excel sheet")
        wb.close()
        return

    names = get_name_to_col(FILE_PATH, SHEET_NAME)

    available_people = []
    for name, col in names.items():
        v = ws.cell(row=row_index, column=col).value
        if v is None or v == "":
            available_people.append(name)

    st.subheader("People who have NOT claimed OT")

    if not available_people:
        st.success("All people have already claimed OT for this date")
        wb.close()
        return

    st.info(f"Bill Amount: ₹{bill_amount:.0f} → You must select **{required_people}** people")

    if required_people > len(available_people):
        st.warning(f"Not enough available people! Required: {required_people}, Available: {len(available_people)}")

    selected_people = st.multiselect(
        f"Select {required_people} people",
        available_people,
        max_selections=required_people,
        key="ot_selected_people"
    )

    if st.button("Submit OT Claim", key="submit_ot_claim"):
        if required_people > 0 and len(selected_people) != min(required_people, len(available_people)):
            st.error("Please select the required number of people.")
            wb.close()
            return

        for person in selected_people:
            ws.cell(row=row_index, column=names[person]).value = "OT"

        try:
            wb.save(FILE_PATH)
            st.success("OT Tracker Updated!!")
            st.info(f"You can Copy names : {'; '.join(selected_people)}")
        except Exception as e:
            st.error(f"Failed to save workbook: {e}")

        wb.close()
        return

    wb.close()

def render_names_finder_tab():
    st.subheader("Names Finder")

    nf_date = st.date_input("Select Date", value=date.today(), key="nf_date")

    try:
        wb = load_workbook(FILE_PATH, data_only=True)
    except Exception as e:
        st.error(f"Unable to open {FILE_PATH}: {e}")
        return

    if SHEET_NAME not in wb.sheetnames:
        st.error(f"Sheet '{SHEET_NAME}' not found in {FILE_PATH}")
        wb.close()
        return

    ws = wb[SHEET_NAME]
    row_index = find_row_for_date(ws, nf_date)
    if row_index is None:
        st.error("Date not found in Excel sheet")
        wb.close()
        return

    names = get_name_to_col(FILE_PATH, SHEET_NAME)

    claimed, unclaimed = [], []
    for name, col in names.items():
        v = ws.cell(row=row_index, column=col).value
        if v is None or v == "":
            unclaimed.append(name)
        else:
            claimed.append(name)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("### ✅ Claimed")
        st.write(claimed if claimed else "No one has claimed yet.")
    with c2:
        st.markdown("### ⬜ Not Claimed")
        st.write(unclaimed if unclaimed else "Everyone has claimed.")

    wb.close()

def render_bills_repo_tab():
    st.subheader("Bills Repo")

    ensure_bills_storage()
    df = read_bills_index()

    # Load names from Excel header for recommendations
    all_names = get_all_names(FILE_PATH, SHEET_NAME)
    all_names = [name for name in all_names if name != "Day"]  
    if not all_names:
        st.error("No names found in Excel header row (Row 1, from Column B onwards).")
        return

    # -------- Find bills --------
    st.markdown("### Find Previously Uploaded Bills")

    col3, col2 = st.columns([1, 1])
    # with col2:
    #     pass
    #     # filter_date = st.date_input("Bill Date", value=date.today(), key="bill_filter_date")
    with col3:
        filter_name = st.selectbox(
            "Filter by name",
            options=all_names,
            index=0,
            key="bill_filter_name_select"
        )

    df_view = df.copy()
    if not df_view.empty:
        if filter_name != "All":
            df_view = df_view[df_view["user_name"].astype(str).str.lower() == filter_name.lower()]

    if df_view.empty:
        st.info("No bills found for the selected filter.")
    else:
        st.dataframe(df_view.sort_values(by="uploaded_at", ascending=False), use_container_width=True)

        # Optional: download buttons
        st.markdown("#### Download (latest 10 shown)")
        for i, row in df_view.sort_values(by="uploaded_at", ascending=False).head(10).iterrows():
            path = str(row["stored_path"])
            if os.path.exists(path):
                with open(path, "rb") as f:
                    st.download_button(
                        label=f"Download: {row['file_name']} ({row['user_name']} - {row['ot_date']})",
                        data=f,
                        file_name=os.path.basename(path),
                        mime="application/octet-stream",
                        key=f"dl_{i}"
                    )

    st.divider()

    # -------- Upload bill --------
    st.markdown("### Upload a Bill")

    upload_date = st.date_input("OT Date for this bill", value=date.today(), key="bill_upload_date")
    upload_name = st.selectbox(
        "User Name",
        options=all_names,
        index=0,
        key="bill_upload_name_select"
    )
    uploaded_file = st.file_uploader(
        "Choose bill file (pdf/jpg/png)",
        type=["pdf", "jpg", "jpeg", "png"],
        key="bill_uploader"
    )

    if st.button("Save Bill", key="bill_save_btn"):
        if uploaded_file is None:
            st.error("Please choose a file to upload.")
            return

        safe_name = sanitize_filename(upload_name)
        safe_orig = sanitize_filename(uploaded_file.name)

        # Store in Bills/YYYY-MM-DD/username__originalname.ext
        date_folder = os.path.join(BILLS_DIR, upload_date.isoformat())
        os.makedirs(date_folder, exist_ok=True)

        stored_filename = f"{safe_name}__{safe_orig}"
        stored_path = os.path.join(date_folder, stored_filename)

        # Avoid overwrite by adding counter
        base, ext = os.path.splitext(stored_filename)
        counter = 1
        while os.path.exists(stored_path):
            stored_filename = f"{base}__{counter}{ext}"
            stored_path = os.path.join(date_folder, stored_filename)
            counter += 1

        try:
            with open(stored_path, "wb") as f:
                f.write(uploaded_file.getbuffer())

            append_bill_index(upload_date, upload_name, uploaded_file.name, stored_path)

            display_path = stored_path.replace("\\", "/")
            st.success("Bill saved successfully!")
            st.info(f"Saved to: {display_path}")

            st.rerun()
        except Exception as e:
            st.error(f"Failed to save bill: {e}")

# -----------------------------
# MAIN UI
# -----------------------------
tab_ot, tab_bills = st.tabs(["OT Meal Tracker", "Bills Repo"])

with tab_ot:
    render_ot_tab()

# with tab_names:
#     render_names_finder_tab()

with tab_bills:
    render_bills_repo_tab()
